"""v3 -> v4: tidy the KU support fields and attach contact emails.

Two changes only:

1. "KU faculty focus" drops the "Alle" family of values. Saying a programme is open
   to every faculty tells a reader nothing, so the column now holds a real faculty
   list (SCIENCE / SUND) or nothing at all.
2. New "KU contact email" column. Addresses come from the KU LH preaward deck:
   lighthouse@ku.dk (slide 62), POC@adm.ku.dk (slides 43-44) and the three Research
   Funding Support campus inboxes (slide 63). Preaward is organised by campus rather
   than faculty, so those rows carry all three and the reader picks their own campus.
   Multiple contacts are stored as "Label: address" pairs separated by " | ".
"""

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
SRC = BASE / "funds with KU support - v3.csv"
OUT_CSV = BASE / "funds with KU support - v4.csv"
OUT_XLSX = BASE / "funds with KU support - v4.xlsx"

EMAIL_COL = "KU contact email"

LIGHTHOUSE_EMAIL = "lighthouse@ku.dk"
POC_EMAIL = "POC@adm.ku.dk"
PREAWARD_EMAILS = " | ".join(
    [
        "Frederiksberg: forskningsfinansiering-frbplus@adm.ku.dk",
        "Nørre: forskningsfinansiering-noerre@adm.ku.dk",
        "Søndre: forskningsfinansiering-soendre@adm.ku.dk",
    ]
)

# Programmes with their own inbox in the deck; these win over the unit address.
PROGRAMME_EMAILS = {
    "UCPH Proof of Concept Fund (POC MAX)": POC_EMAIL,
    "UCPH Proof of Concept Fund (POC-TO-GO)": POC_EMAIL,
}

FACULTIES = ["SCIENCE", "SUND"]


def clean_faculty(value: str) -> str:
    """Keep named faculties; drop 'Alle' and any parenthetical qualifiers."""
    found = [f for f in FACULTIES if re.search(rf"\b{f}\b", value, re.IGNORECASE)]
    return ", ".join(found)


def contact_email(row: dict[str, str]) -> str:
    if row["Name"] in PROGRAMME_EMAILS:
        return PROGRAMME_EMAILS[row["Name"]]
    unit = row.get("KU support unit", "").strip()
    if unit == "Lighthouse":
        return LIGHTHOUSE_EMAIL
    if unit == "Preaward":
        return PREAWARD_EMAILS
    return ""


def main() -> None:
    with SRC.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        headers = list(reader.fieldnames or [])
        rows = [{k: (v or "").strip() for k, v in r.items()} for r in reader]

    headers.insert(headers.index("KU contact hint"), EMAIL_COL)

    dropped: list[str] = []
    for row in rows:
        before = row.get("KU faculty focus", "")
        after = clean_faculty(before)
        if before != after:
            dropped.append(f"{row['Name']}: {before!r} -> {after!r}")
        row["KU faculty focus"] = after
        row[EMAIL_COL] = contact_email(row)

    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter=";", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)

    export_xlsx(headers, rows)

    with_email = sum(1 for r in rows if r[EMAIL_COL])
    no_email = [
        r["Name"] for r in rows
        if r.get("KU support unit") not in ("", "—") and not r[EMAIL_COL]
    ]
    print(f"Faculty values rewritten: {len(dropped)}")
    for line in dropped:
        print("  " + line)
    print(f"\nRows with a KU contact email: {with_email}")
    print(f"KU-supported rows still without one ({len(no_email)}): {', '.join(no_email)}")
    print(f"\nRows: {len(rows)}  ->  {OUT_CSV.name} / {OUT_XLSX.name}")


def export_xlsx(headers: list[str], rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Funds"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            val = row.get(h, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if h == "Link" and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
            # Only a lone address can be a hyperlink; campus sets stay plain text.
            if h == EMAIL_COL and val and "|" not in val:
                cell.hyperlink = f"mailto:{val}"
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    widths = [18, 28, 35, 40, 22, 18, 10, 18, 16, 55, 18, 22, 20, 20, 24, 22, 40]
    for i, w in enumerate(widths[: len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
